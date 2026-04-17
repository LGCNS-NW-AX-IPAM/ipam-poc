import base64
import json
import os
import re
import smtplib
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from io import BytesIO
from typing import Dict, List, Optional, Tuple

from dotenv import load_dotenv
from langchain_core.messages import AIMessage, BaseMessage, HumanMessage, SystemMessage
from langchain_google_genai import ChatGoogleGenerativeAI
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.client.ntoss_client import NtossClient
from app.repositories.candidate.candidate_repository import CandidateRepository


class CandidateService:
    RECLAIM_REASON_COL = "회수 사유"

    REQUIRED_HEADERS = [
        "DHCP Server IP",
        "IP블록",
        "인프라팀",
        "네트워크 이름",
        "네트워크 ID",
        "Primary 여부",
        "사용률(%)",
    ]

    def __init__(self):
        load_dotenv()
        self.ntoss = NtossClient()
        self.llm = ChatGoogleGenerativeAI(
            model="gemini-2.5-flash",
            temperature=0,
            google_api_key=os.getenv("GOOGLE_API_KEY"),
        )
        self._classification_cache: Dict[str, bool] = {}

    @staticmethod
    def _convert_to_messages(messages: List[dict]) -> List[BaseMessage]:
        converted: List[BaseMessage] = []
        for message in messages:
            role = message.get("role", "user")
            content = message.get("content", "")
            converted.append(HumanMessage(content=content) if role == "user" else AIMessage(content=content))
        return converted

    def infer_upload_mode_from_history(self, history: List[dict]) -> str:
        print("\n🚀 [FUNC: infer_upload_mode_from_history(CandidateService)]")
        if not history:
            print("ℹ️ history 없음 -> mode=extract")
            return "extract"
        print(f"🎯 분석할 대화 기록(Candidate): {history}")

        system_prompt = """
        당신은 IP 회수 후보 업로드 모드 판별기입니다.
        최근 대화 기록을 보고 아래 중 하나로만 분류하세요.

        [모드 종류]
        - EXTRACT : 후보 추출 단계
        - FINALIZE: 후보 확정(DB 반영) 단계

        [판단 기준]
        - 가장 최근에 한 대화에 "확정", "반영" 등 후보 확정과 관련된 단어가 포함되어 있으면 FINALIZE 모드로 판단
        - 그 외 모든 경우는 EXTRACT 모드로 판단

        [출력 규칙]
        - 반드시 라벨 하나만 출력: EXTRACT / FINALIZE
        - 설명 없이 한 단어만 출력
        """
        try:
            # NOTE:
            # history 자체를 메시지 배열로 그대로 주고 마지막 turn이 AIMessage로 끝나면,
            # 일부 경우 모델이 다음 사용자 지시를 받지 못한 것으로 해석해 빈 content를 반환할 수 있습니다.
            # 따라서 대화 이력을 문자열 컨텍스트로 넘기고, 마지막에 명시적인 분류 요청 HumanMessage를 붙입니다.
            transcript_lines = []
            for item in history:
                role = str(item.get("role", "user")).upper()
                content = str(item.get("content", "")).strip()
                transcript_lines.append(f"{role}: {content}")
            transcript_text = "\n".join(transcript_lines)
            classification_request = (
                "아래 대화 이력을 보고 업로드 모드를 분류하세요.\n"
                f"{transcript_text}\n\n"
                "정답 라벨만 출력: EXTRACT 또는 FINALIZE"
            )

            res = self.llm.invoke(
                [
                    SystemMessage(content=system_prompt),
                    HumanMessage(content=classification_request),
                ]
            )
            print(f"🎯 분석 결과 res (Candidate): {res}")
            raw = str(res.content or "").strip().upper()
            print(f"🎯 분석된 모드 raw (Candidate): {raw}")
            if not raw:
                raise ValueError(
                    "LLM returned empty content. "
                    f"finish_reason={res.response_metadata.get('finish_reason') if hasattr(res, 'response_metadata') else 'unknown'}"
                )
            intent = next(
                (x for x in ["FINALIZE", "EXTRACT"] if x in raw),
                "EXTRACT",
            )
            mode = "finalize" if intent == "FINALIZE" else "extract"
            print(f"🧭 inferred intent={intent} -> mode={mode}")
            return mode
        except Exception as e:
            print(f"❌ infer_upload_mode_from_history error: {e}")
            print("↪ fallback mode=extract")
            return "extract"

    @staticmethod
    def _load_team_email_map() -> Dict[str, str]:
        raw = os.getenv("INFRA_TEAM_EMAIL_MAP", "").strip()
        if not raw:
            return {}
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                return {str(k).strip(): str(v).strip() for k, v in parsed.items() if str(v).strip()}
        except Exception:
            pass
        mapping: Dict[str, str] = {}
        for pair in raw.split(";"):
            if ":" not in pair:
                continue
            key, value = pair.split(":", 1)
            if key.strip() and value.strip():
                mapping[key.strip()] = value.strip()
        return mapping

    @staticmethod
    def _build_review_excel_bytes(selected_ips: List[dict]) -> bytes:
        wb = Workbook()
        ws = wb.active
        reason_col = CandidateService.RECLAIM_REASON_COL
        with_snap = [x for x in selected_ips if x.get("excel_row")]
        if with_snap:
            hdrs = list(with_snap[0]["excel_row"].keys())
            if reason_col not in hdrs:
                hdrs = hdrs + [reason_col]
            ws.append(hdrs)
            for item in with_snap:
                er = dict(item["excel_row"])
                er[reason_col] = str(item.get("decision_reason", "") or "")
                ws.append([er.get(h) for h in hdrs])
        else:
            ws.append(["네트워크 ID", "IP블록", "인프라팀", "담당 이메일", reason_col])
            for item in selected_ips:
                ws.append(
                    [
                        item.get("nw_id"),
                        item.get("ip_address"),
                        item.get("owner_team"),
                        item.get("owner_email"),
                        str(item.get("decision_reason", "") or ""),
                    ]
                )
        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()

    @staticmethod
    def build_review_excel_base64(selected_ips: List[dict]) -> Optional[str]:
        if not selected_ips:
            return None
        return base64.b64encode(CandidateService._build_review_excel_bytes(selected_ips)).decode("ascii")

    def send_review_mails(self, selected_ips: List[dict], override_recipients: Optional[List[str]] = None) -> Dict:
        print("\n🚀 [FUNC: send_review_mails(CandidateService)]")
        if not selected_ips:
            print("⚠️ selected_ips is empty")
            return {"sent_count": 0, "failed": []}

        if override_recipients:
            recipients = sorted({str(email).strip() for email in override_recipients if str(email).strip()})
        else:
            team_email_map = self._load_team_email_map()
            default_email = os.getenv("CANDIDATE_DEFAULT_OWNER_EMAIL", "no-reply@ipam.local")
            recipients = sorted(
                {
                    team_email_map.get(item.get("owner_team", "").strip()) or item.get("owner_email") or default_email
                    for item in selected_ips
                    if item.get("owner_team") or item.get("owner_email")
                }
            )
        print(f"📨 recipients={recipients}")
        gmail_user = os.getenv("GMAIL_USER")
        raw_gmail_password = os.getenv("GMAIL_APP_PASSWORD", "")
        # App Password는 종종 "abcd efgh ijkl mnop" 형태로 저장되어 공백 제거가 필요합니다.
        gmail_password = raw_gmail_password.replace(" ", "").strip()
        print(f"🔐 gmail_user={'set' if gmail_user else 'missing'}")
        print(f"🔐 gmail_app_password={'set' if gmail_password else 'missing'} (len={len(gmail_password)})")
        subject = "[IPAM] IP 회수 후보 검토 요청"

        body_lines = [
            "안녕하세요. IPAM AI Agent입니다.",
            "아래 IP 회수 후보에 대한 검토를 요청드립니다.",
            "첨부 엑셀은 선정된 회수 후보만 포함합니다(제외 행 제거).",
            "각 행의 '회수 사유' 열에 선정 근거가 기재되어 있습니다.",
            "",
        ]
        for item in selected_ips[:30]:
            reason = str(item.get("decision_reason", "") or "").strip()
            reason_part = f" | 회수 사유: {reason}" if reason else ""
            body_lines.append(
                f"- {item.get('owner_team')} | {item.get('nw_id')} | {item.get('ip_address')}{reason_part}"
            )
        body_lines.append("")
        body_lines.append("검토 후 회신 부탁드립니다.")
        body = "\n".join(body_lines)

        if not recipients:
            print("⚠️ recipients is empty after resolution")
            return {"sent_count": 0, "failed": ["NO_RECIPIENT"]}

        if not gmail_user or not gmail_password:
            print("⚠️ gmail credentials missing -> mock success")
            return {"sent_count": len(recipients), "failed": []}

        xlsx_bytes = self._build_review_excel_bytes(selected_ips)
        attach_name = "ip_reclaim_candidates_review.xlsx"
        failed = []
        failed_reasons: Dict[str, str] = {}
        for to_email in recipients:
            print(f"➡️ start send to={to_email}")
            msg = MIMEMultipart()
            msg["Subject"] = subject
            msg["From"] = gmail_user
            msg["To"] = to_email
            msg.attach(MIMEText(body, _charset="utf-8"))
            part = MIMEApplication(
                xlsx_bytes,
                _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            part.add_header("Content-Disposition", "attachment", filename=attach_name)
            msg.attach(part)
            try:
                print("   - connect smtp.gmail.com:465")
                with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
                    print("   - login")
                    smtp.login(gmail_user, gmail_password)
                    print("   - send_message")
                    smtp.send_message(msg)
                print(f"✅ sent to={to_email}")
            except Exception:
                import traceback
                err = traceback.format_exc()
                print(f"❌ send failed to={to_email}\n{err}")
                failed.append(to_email)
                failed_reasons[to_email] = err
        print(f"📊 mail result sent={len(recipients) - len(failed)} failed={len(failed)}")
        return {
            "sent_count": len(recipients) - len(failed),
            "failed": failed,
            "failed_reasons": failed_reasons,
        }

    @staticmethod
    def _normalize_header(value) -> str:
        return str(value).strip() if value is not None else ""

    @staticmethod
    def _to_percent(value) -> float:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value) * 100 if float(value) <= 1 else float(value)
        text = str(value).strip().replace("%", "")
        try:
            return float(text)
        except ValueError:
            return 0.0

    @staticmethod
    def _is_non_primary(value) -> bool:
        return str(value).strip().upper() != "Y" if value is not None else True

    def _is_accommodation_by_llm(self, name: str) -> bool:
        normalized = (name or "").strip()
        if not normalized:
            return False
        if normalized in self._classification_cache:
            return self._classification_cache[normalized]

        prompt = f"""
        아래 명칭이 '특정 요일/시점에 사용량이 급증·급감할 수 있는 단기 숙박 시설'인지 분류하세요.
        단기 숙박 시설 예: 기숙사, 호텔, 숙박업소, 모텔, 리조트, 게스트하우스 등
        일반 주거 아파트/일반 상업시설은 KEEP으로 분류하세요.
        명칭: "{normalized}"
        출력은 EXCLUDE 또는 KEEP 한 단어만 반환하세요.
        """
        try:
            response = self.llm.invoke(prompt)
            is_excluded = "EXCLUDE" in str(response.content).upper()
        except Exception:
            is_excluded = False
        self._classification_cache[normalized] = is_excluded
        return is_excluded

    def _llm_generate_reason(self, row_context: Dict, excluded: bool) -> str:
        prompt = f"""
        아래 판정 결과를 관리자에게 설명할 한 줄 사유를 작성하세요.
        - excluded={excluded}
        - 판정 기준: 사용률 임계치, Primary 여부, 숙소형 시설 제외
        - 데이터: {row_context}
        40자 이내 한국어 문장으로 출력하세요.
        """
        try:
            res = self.llm.invoke(prompt)
            return str(res.content).strip()
        except Exception:
            return "정책 기준에 따라 자동 판정됨"

    def extract_candidates_from_excel(
        self,
        db: Session,
        file_bytes: bytes,
        extraction_batch_id: str,
        usage_threshold: float,
        default_owner_email: str,
    ) -> Dict:
        _ = db
        wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("엑셀 파일이 비어 있습니다.")

        headers = [self._normalize_header(h) for h in rows[0]]
        header_index = {name: idx for idx, name in enumerate(headers)}
        missing_headers = [h for h in self.REQUIRED_HEADERS if h not in header_index]
        if missing_headers:
            raise ValueError(f"필수 컬럼이 없습니다: {', '.join(missing_headers)}")

        inserted = 0
        skipped = 0
        excluded_by_accommodation = 0
        selected_preview: List[Dict] = []
        excluded_details: List[Dict] = []
        selected_ips: List[Dict] = []
        seen_keys: set[Tuple[str, str]] = set()

        for row in rows[1:]:
            dhcp_ip = row[header_index["DHCP Server IP"]]
            ip_block = str(row[header_index["IP블록"]] or "").strip()
            owner_team = row[header_index["인프라팀"]]
            network_name = str(row[header_index["네트워크 이름"]] or "")
            nw_id = str(row[header_index["네트워크 ID"]] or "")
            primary_flag = row[header_index["Primary 여부"]]
            usage_raw = row[header_index["사용률(%)"]]
            display_ip = ip_block if ip_block else f"{dhcp_ip}/32"

            if not dhcp_ip or not nw_id or not owner_team:
                skipped += 1
                continue

            usage_percent = self._to_percent(usage_raw)
            ntoss_result = self.ntoss.get_apartment_info_by_nw_id(nw_id=nw_id)
            apartment_name = str(ntoss_result.get("apartment_name", "") or "")
            network_is_accommodation = self._is_accommodation_by_llm(network_name)
            apartment_is_accommodation = self._is_accommodation_by_llm(apartment_name)

            row_context = {
                "nw_id": nw_id,
                "ip_address": display_ip,
                "owner_team": str(owner_team),
                "usage_percent": usage_percent,
                "is_non_primary": self._is_non_primary(primary_flag),
                "network_name": network_name,
                "apartment_name": apartment_name,
                "network_name_is_accommodation": network_is_accommodation,
                "apartment_name_is_accommodation": apartment_is_accommodation,
            }
            is_under_threshold = usage_percent < usage_threshold
            is_non_primary = self._is_non_primary(primary_flag)
            is_excluded_accommodation = network_is_accommodation or apartment_is_accommodation
            should_select = is_under_threshold and is_non_primary and (not is_excluded_accommodation)

            if not should_select:
                skipped += 1
                if is_excluded_accommodation:
                    excluded_by_accommodation += 1
                reasons = []
                if not is_under_threshold:
                    reasons.append(f"사용률 {usage_percent:.2f}%가 기준({usage_threshold:.2f}%) 미만이 아님")
                if not is_non_primary:
                    reasons.append("Primary 여부가 Y이므로 제외")
                if is_excluded_accommodation:
                    reasons.append("네트워크명 또는 NTOSS 아파트명이 숙소형 시설로 분류됨")
                excluded_details.append(
                    {
                        "nw_id": nw_id,
                        "ip_address": display_ip,
                        "owner_team": str(owner_team),
                        "usage_percent": usage_percent,
                        "network_name": network_name,
                        "apartment_name": apartment_name,
                        "exclude_reason": " / ".join(reasons) if reasons else "정책 기준 미충족",
                    }
                )
                continue

            unique_key = (nw_id, display_ip)
            if unique_key in seen_keys:
                skipped += 1
                excluded_details.append(
                    {
                        "nw_id": nw_id,
                        "ip_address": display_ip,
                        "owner_team": str(owner_team),
                        "usage_percent": usage_percent,
                        "network_name": network_name,
                        "apartment_name": apartment_name,
                        "exclude_reason": "엑셀 내 중복 대상",
                    }
                )
                continue

            seen_keys.add(unique_key)
            inserted += 1
            excel_row = {h: row[header_index[h]] if header_index[h] < len(row) else None for h in headers}
            selected_item = {
                "nw_id": nw_id,
                "ip_address": display_ip,
                "owner_team": str(owner_team),
                "owner_email": default_owner_email,
                "usage_percent": usage_percent,
                "network_name": network_name,
                "apartment_name": apartment_name,
                "decision_reason": self._llm_generate_reason(row_context, excluded=False),
            }
            selected_ips.append(
                {
                    "nw_id": selected_item["nw_id"],
                    "ip_address": selected_item["ip_address"],
                    "owner_team": selected_item["owner_team"],
                    "owner_email": selected_item["owner_email"],
                    "decision_reason": selected_item["decision_reason"],
                    "excel_row": excel_row,
                }
            )
            selected_preview.append(selected_item)

        return {
            "batch_id": extraction_batch_id,
            "usage_threshold": usage_threshold,
            "selected_count": inserted,
            "skipped_count": skipped,
            "excluded_by_accommodation_count": excluded_by_accommodation,
            "selection_policy": {
                "usage_threshold_percent": usage_threshold,
                "non_primary_required": True,
                "exclude_accommodation": True,
            },
            "selected_preview": selected_preview,
            "excluded_details": excluded_details,
            "selected_ips": selected_ips,
            "requires_finalize": True,
        }

    def _insert_confirmed_candidates(self, db: Session, selected_ips: List[dict], extraction_batch_id: str = "") -> Dict:
        if not selected_ips:
            return {"inserted_count": 0, "skipped_count": 0}
        normalized = []
        fallback_email = os.getenv("CANDIDATE_DEFAULT_OWNER_EMAIL", "no-reply@ipam.local")
        for item in selected_ips:
            copied = dict(item)
            copied["owner_email"] = str(copied.get("owner_email", "")).strip() or fallback_email
            normalized.append(copied)
        repo = CandidateRepository(db)
        return repo.insert_confirmed_candidates(normalized, extraction_batch_id)

    def finalize_candidates_from_excel(
        self,
        db: Session,
        file_bytes: bytes,
        extraction_batch_id: str,
        usage_threshold: float,
        default_owner_email: str,
    ) -> Dict:
        _ = usage_threshold
        wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("엑셀 파일이 비어 있습니다.")

        headers = [self._normalize_header(h) for h in rows[0]]
        header_index = {name: idx for idx, name in enumerate(headers)}
        missing_headers = [h for h in self.REQUIRED_HEADERS if h not in header_index]
        if missing_headers:
            raise ValueError(f"필수 컬럼이 없습니다: {', '.join(missing_headers)}")

        prepared_ips: List[Dict] = []
        finalized_preview: List[Dict] = []
        for row in rows[1:]:
            dhcp_ip = row[header_index["DHCP Server IP"]]
            ip_block = str(row[header_index["IP블록"]] or "").strip()
            owner_team = str(row[header_index["인프라팀"]] or "").strip()
            nw_id = str(row[header_index["네트워크 ID"]] or "").strip()
            usage_raw = row[header_index["사용률(%)"]]
            display_ip = ip_block if ip_block else f"{dhcp_ip}/32"
            if not dhcp_ip or not nw_id or not owner_team:
                continue
            prepared_ips.append(
                {
                    "nw_id": nw_id,
                    "ip_address": display_ip,
                    "owner_team": owner_team,
                    "owner_email": default_owner_email,
                }
            )
            finalized_preview.append(
                {
                    "owner_team": owner_team,
                    "nw_id": nw_id,
                    "ip_address": display_ip,
                    "usage_percent": self._to_percent(usage_raw),
                }
            )

        insert_result = self._insert_confirmed_candidates(
            db=db,
            selected_ips=prepared_ips,
            extraction_batch_id=extraction_batch_id,
        )
        return {
            "batch_id": extraction_batch_id,
            "selected_count": len(prepared_ips),
            "excluded_by_accommodation_count": 0,
            "inserted_count": insert_result.get("inserted_count", 0),
            "skipped_count": insert_result.get("skipped_count", 0),
            "finalized_preview": finalized_preview,
        }

    def build_extract_response_message(self, result: Dict) -> str:
        prompt = f"""
        당신은 IPAM AI Assistant입니다.
        아래 데이터를 바탕으로 "정해진 양식"으로만 응답하세요.

        [중요 규칙]
        1) 후보 목록(selected_preview)과 제외 목록(excluded_details)을 절대 요약/생략하지 말고 전부 출력하세요.
        2) 기준 IP사용률(usage_threshold_percent)을 반드시 명시하세요.
        3) 제외 목록은 각 항목의 exclude_reason을 그대로 포함하세요.
        4) 데이터에 없는 내용을 임의로 만들지 마세요.
        5) 아래 출력 템플릿의 제목/순서를 그대로 지키세요.
        6) 마지막 안내 문장은 반드시 아래 문장과 100% 동일해야 합니다.
           후보 확인 후 '메일 발송'이라고 입력하면 검토 메일을 인프라 담당자에게 발송하고, 수정이 필요하다면 수정할 내용을 입력해주세요.

        [출력 템플릿]
        엑셀 분석 결과 요약
        - 후보 건수: {{selected_count}}건
        - 제외 건수: {{skipped_count}}건
        - 기준 IP사용률: {{usage_threshold_percent}}%
        - 선정 기준: 사용률 미달 + Non-primary + 단기 숙박 시설 제외

        후보 목록
        - {{owner_team}} | {{nw_id}} | {{ip_address}} | 사용률 {{usage_percent}}% | 근거: {{decision_reason}}
        - ... (selected_preview의 모든 항목)

        제외 목록
        - {{owner_team}} | {{nw_id}} | {{ip_address}} | 사용률 {{usage_percent}}% | 제외 사유: {{exclude_reason}}
        - ... (excluded_details의 모든 항목)

        후보 확인 후 '메일 발송'이라고 입력하면 검토 메일을 인프라 담당자에게 발송하고, 수정이 필요하다면 수정할 내용을 입력해주세요.

        [입력 데이터]
        {result}
        """
        try:
            res = self.llm.invoke([HumanMessage(content=prompt)])
            return str(res.content).strip()
        except Exception:
            selected_preview = result.get("selected_preview", []) or []
            excluded_details = result.get("excluded_details", []) or []
            usage_threshold = (result.get("selection_policy", {}) or {}).get("usage_threshold_percent", "-")

            lines = [
                "엑셀 분석 결과 요약",
                f"- 후보 건수: {result.get('selected_count', 0)}건",
                f"- 제외 건수: {result.get('skipped_count', 0)}건",
                f"- 기준 IP사용률: {usage_threshold}%",
                "- 선정 기준: 사용률 미달 + Non-primary + 단기 숙박 시설 제외",
                "",
                "후보 목록",
            ]

            if selected_preview:
                for item in selected_preview:
                    lines.append(
                        f"- {item.get('owner_team')} | {item.get('nw_id')} | {item.get('ip_address')} | "
                        f"사용률 {item.get('usage_percent')}% | 근거: {item.get('decision_reason', '정책 기준 충족')}"
                    )
            else:
                lines.append("- 후보 없음")

            lines.append("")
            lines.append("제외 목록")
            if excluded_details:
                for item in excluded_details:
                    lines.append(
                        f"- {item.get('owner_team')} | {item.get('nw_id')} | {item.get('ip_address')} | "
                        f"사용률 {item.get('usage_percent')}% | 제외 사유: {item.get('exclude_reason', '정책 기준 미충족')}"
                    )
            else:
                lines.append("- 제외 없음")

            lines.append("")
            lines.append(
                "후보 확인 후 '메일 발송'이라고 입력하면 검토 메일을 인프라 담당자에게 발송하고, 수정이 필요하다면 수정할 내용을 입력해주세요."
            )
            return "\n".join(lines)

    def build_finalize_response_message(self, result: Dict) -> str:
        selected_count = result.get("selected_count", 0)
        inserted_count = result.get("inserted_count", 0)
        skipped_count = result.get("skipped_count", 0)
        finalized_preview = result.get("finalized_preview", []) or []

        lines = [
            f"회수 후보 확정이 완료되었습니다. 총 {selected_count}건 중 {inserted_count}건이 DB에 반영되었고, 제외/건너뜀은 {skipped_count}건입니다.",
            "",
            "확정된 항목",
        ]

        if finalized_preview:
            for item in finalized_preview:
                lines.append(
                    f"- {item.get('owner_team')} | {item.get('nw_id')} | {item.get('ip_address')} | 사용률 {item.get('usage_percent')}%"
                )
        else:
            lines.append("- 확정된 항목 없음")

        return "\n".join(lines)
